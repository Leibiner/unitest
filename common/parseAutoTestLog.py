# -*- coding: UTF-8 -*-

import os
import sys
import time
from openpyxl import load_workbook
from ftpSync import ftp
from .testResultCommon import TestResultCommon
from sendResultByMail import Mail, HtmlReport
from . import globalVariables
from . import commonBase as com
import urllib.request, urllib.parse, urllib.error
import traceback
import copy
import shutil
from .network import network


class ParseAutoTestLog(TestResultCommon):
    reportUrl = ""
    ftpServer = "http://xxx.xxx.xx.xxx"
    cmdParas = {}
    parasList = ["rFile", "projectKey", "runType", "caseFilePath", "runEnv", "reportCategory", "onlyUploadLog",
                 "mailSubject", "mailSender", "mailList", "sendMail", "mailProjectKey", "showUrlInMail"]
    report_category = 0

    def loadCommandParas(self, argv):
        '''
        :param argv: argument from command line
        :return: True if load argument success from command line, otherwise return False
        '''
        self.cmdParas.clear()
        for i in range(1, len(argv)):
            if argv[i].startswith("-"):
                args = argv[i][1:].split('=')
                if args[0] not in self.parasList:
                    self.showHelpInfo()
                    return False
                self.cmdParas[args[0]] = args[1]
            else:
                self.showHelpInfo()
                return False
        return True

    def showHelpInfo(self):
        '''
        :return: the help Info to use sendReport plugin
        '''
        print("Parameters supported as below:")
        print("-rFile           the result file")
        print("-projectKey      the key to related to the special project")
        print("-caseFilePath    the file paths to update the case description which separate by ,")
        print("-runEnv          the run environment number, 1(QA),2(YZ),3(PROD),5(PRE)")
        print("-runType         the run type, api or ui")
        print("-reportCategory  the report category, 0(regression), 1(xunjiang)")
        print("-onlyUploadLog   only upload the log, True/False")
        print("\nBelow the parameters for send mail")
        print("-sendMail        whether to send mail with the test result, True/False")
        print("-mailProjectKey  the key to related to special project to get the mail list")
        print("-mailSubject     the mail subject in result mail")
        print("-mailSender      the mail address of sender")
        print("-mailList        the mail receivers list")
        print("-showUrlInMail   show the urls are invoked in the test case")

    def logToDb(self, resultXmlName, run_env=""):
        '''
        :param resultXmlName: xml file after nosetests run test
        :param run_env: which env the nosetests runs
        :return: testRsFlag: the test is success or failure
                 reportId: the id on qa platform
        '''
        try:
            startTime = self.getStartTime()
            if run_env == "":
                subjectEnv = self.getRunEnv()
            else:
                subjectEnv = run_env
            projectId, projectName = self.getProjectIdAndName(self.projectKey)
            print("parse test XML report...")
            root = self.readXmlLog(resultXmlName)
            allTests = int(root.get("tests"))
            errors = int(root.get("errors"))
            failures = int(root.get("failures"))
            skip = int(root.get("skip"))
            passCount = allTests - failures - errors - skip
            print("total: %s  pass: %s  fail: %s  error:%s" % (allTests, passCount, failures, errors))

            testRsFlag = True
            if failures != 0 or errors != 0:
                testRsFlag = False

            rule = r"--------------------"
            sumTime = 0.0
            caseNameList = []
            caseRsList = []
            pkgName = ""

            for child in root:
                caseTime = float(child.get("time"))
                sumTime = sumTime + caseTime
                className = child.get("classname")
                if className == r"XXX.testcase":
                    moduleName = className
                    pkgName = "testcase"
                elif "nose.suite" in className:
                    moduleName = "nose.suite"
                    pkgName = "caseError"
                else:
                    moduleName = '.'.join(className.split('.')[0:-1])
                    # if className.find("testcase.") > -1:
                    #     moduleName = className.split("testcase.")[1]
                    pkgName = moduleName.split(".")[0]

                caseName = child.get("name")

                # remove (..) when run using parameterize
                pos = caseName.find('(')
                if pos > -1:
                    caseName = caseName[:pos]

                caseNameList.append(caseName)
                # resCase = 1 as pass, 2 as failure,3 as error,4 as skip
                resCase = 1
                caseNote = ""
                caseSysOut = ""
                msg = ""

                for subNode in child:
                    caseStatus = subNode.tag
                    if caseStatus == "failure" or caseStatus == "error" or caseStatus == "skipped ":
                        msg = subNode.get("message")
                        if r"'" in msg:
                            msg = msg.replace(r"'", '"')
                        if len(msg) > 60000:
                            msg = msg[:60000]
                        msgList = msg.split(rule)
                        caseNote = msgList[0]

                    if caseStatus == "failure":
                        resCase = 2
                        break
                    elif caseStatus == "error":
                        resCase = 3
                        break
                    elif caseStatus == "skipped":
                        resCase = 4
                        caseNote = subNode.get("message")
                    elif caseStatus == "system-out":
                        caseSysOut = subNode.get("system-out")
                    else:
                        resCase = 1
                caseRsList.append((moduleName, caseName, resCase, caseTime, caseNote, msg, caseSysOut))

            print("insert batch result...")
            data = {'dbName': 'ArReport', 'fields': ['report_id'],
                    'data': [{'project_id': projectId, 'project_key': self.projectKey, 'report_name': projectName,
                              'report_type': subjectEnv, 'case_number': allTests, 'pass_field': passCount,
                              'fail': failures,
                              'error': errors, 'skip': skip, 'run_time': sumTime, 'start_time': startTime,
                              'create_time': com.currentDateTime('%Y-%m-%d %H:%M:%S'),
                              'update_time': com.currentDateTime('%Y-%m-%d %H:%M:%S'),
                              'report_category': self.report_category, 'report_url': self.reportUrl,
                              'execute_ip': network.get_local_ip(), 'execute_channel': 'parseLog'}]}
            reportId = self.sendAddInfoApiRequest(data)[0]['report_id']

            casesList = []
            for case in caseRsList:
                casesList.append(
                    {'report_id': reportId, 'case_module': case[0], 'case_name': case[1], 'result': case[2],
                     'run_time': case[3], 'message': case[4], 'detail': case[5], 'system_out': case[6],
                     'create_time': com.currentDateTime('%Y-%m-%d %H:%M:%S')})
            try:
                data = {'dbName': 'ArResult', 'data': casesList}
                self.sendAddInfoApiRequest(data)
            except Exception as e:
                print("@@@@@@@@@@@logToDb:%s" % str(e))
        except Exception as e:
            print(e.message)
            return None, None
        return testRsFlag, reportId

    def logExcelResultToDb(self, resultExcelName, run_env=""):
        '''
        same as method: logToDb except result source
        :param resultExcelName: the excel result after nosetests run test
        :param run_env: the env which nosetests run
        :return: testRsFlag: the test is success or failure
                 reportId: the id on qa platform
        '''
        try:
            startTime = self.getStartTime()
            if run_env == "":
                subjectEnv = self.getRunEnv()
            else:
                subjectEnv = run_env
            projectId, projectName = self.getProjectIdAndName(self.projectKey)

            wb = load_workbook(resultExcelName)
            # Get the test sheet names
            sheetNames = []
            for sName in wb.get_sheet_names():
                if sName.startswith('test'):
                    sheetNames.append(sName)

            totalTime = 0.0
            caseRsList = []
            iPass = iFail = iSkip = 0
            for sheetName in sheetNames:
                self.loadDataCols(wb, sheetName)
                sheetInfo = wb.get_sheet_by_name(sheetName)

                for iRow in range(2, sheetInfo.max_row + 1):
                    runTime = self.getCellInfo(sheetInfo, "Execute Time", iRow)
                    if runTime is None:
                        runTime = 0
                    runTime = float(runTime)
                    totalTime += runTime
                    data = None
                    className = sheetName
                    caseInfo = self.getCellInfo(sheetInfo, "Info", iRow)
                    resCase = 1
                    status = self.getCellInfo(sheetInfo, "Execute Result", iRow)
                    exception = self.getCellInfo(sheetInfo, "Exception", iRow)
                    if exception is None:
                        exception = ''
                    exception = exception.replace("'", '"')

                    if status == "Fail":
                        iFail += 1
                        resCase = 2
                    elif status == "Pass":
                        iPass += 1
                        resCase = 1
                    else:
                        continue
                        # iSkip += 1
                        # resCase = 4
                    self._insertCaseDescToDb(caseInfo, "", className)
                    caseRsList.append((className, caseInfo, resCase, runTime, exception, exception, ''))

            allTests = iPass + iFail + iSkip
            print("total: %s  pass: %s  fail: %s" % (iPass + iFail + iSkip, iPass, iFail))

            testRsFlag = True
            if iFail != 0:
                testRsFlag = False

            print("insert batch result...")
            data = {'dbName': 'ArReport', 'fields': ['report_id'],
                    'data': [{'project_id': projectId, 'project_key': self.projectKey, 'report_name': projectName,
                              'report_type': subjectEnv, 'case_number': allTests, 'pass_field': iPass,
                              'fail': iFail, 'error': 0, 'skip': iSkip, 'run_time': totalTime, 'start_time': startTime,
                              'create_time': com.currentDateTime('%Y-%m-%d %H:%M:%S'),
                              'update_time': com.currentDateTime('%Y-%m-%d %H:%M:%S'),
                              'report_category': self.report_category}]}
            reportId = self.sendAddInfoApiRequest(data)[0]['report_id']

            casesList = []
            for case in caseRsList:
                casesList.append(
                    {'report_id': reportId, 'case_module': case[0], 'case_name': case[1], 'result': case[2],
                     'run_time': case[3], 'message': case[4], 'detail': case[5], 'system_out': case[6]})
            try:
                data = {'dbName': 'ArResult', 'data': casesList}
                self.sendAddInfoApiRequest(data)
            except Exception as e:
                print("@@@@@@@@@@@logExcelResultToDb:%s" % str(e))
        except Exception as e:
            print(e.message)
            return None, None
        return testRsFlag, reportId

    def logUILogToDb(self, resultXmlName, run_env="", report_category=0):
        '''
        :param resultXmlName: the ui log after nosetests run test
        :param run_env: the env which nosetests run
        :param report_category: the category of report
        :return: testRsFlag: the test is success or failure
                 reportId: the id on qa platform
                 reportURL: the url of report
        '''
        try:
            testRsFlag = True
            startTime = self.getStartTime()
            if run_env == "":
                subjectEnv = self.getRunEnv()
            else:
                subjectEnv = run_env
            projectId, projectName = self.getProjectIdAndName(self.projectKey)
            print("parse test XML report...")
            root = self.readXmlLog(resultXmlName)

            sumTime = 0
            caseNameList = []
            caseRsList = []
            pkgName = self.projectKey
            passCount = 0
            failCount = 0
            suites = root.findall('suite/suite')
            for suite in suites:
                moduleName = suite.get("name")
                suiteStatus = suite.find("status")
                sumTime += self._getDiffTime(suiteStatus.get("endtime"), suiteStatus.get("starttime"))
                for test in suite.findall("test"):
                    status = test.find("status")
                    caseTime = self._getDiffTime(status.get("endtime"), status.get("starttime"))
                    caseName = test.get("name")
                    caseDesp = test.get("comment")
                    self._insertCaseDescToDb(caseName, caseDesp, moduleName)
                    caseNameList.append(caseName)
                    # resCase = 1 as pass, 2 as failure,3 as error,4 as skip
                    resCase = 1
                    caseNote = ""
                    caseSysOut = ""
                    msg = ""

                    caseStatus = status.get("status")
                    if caseStatus == "PASS":
                        resCase = 1
                        passCount += 1
                    elif caseStatus == "FAIL":
                        testRsFlag = False
                        resCase = 2
                        failCount += 1
                        for msg1 in test.findall("kw/msg"):
                            msg += "</br>" + msg1.text
                        msg = msg.replace("'", '"')
                        if len(msg) > 60000:
                            msg = msg[:60000]

                    caseRsList.append((moduleName, caseName, resCase, caseTime, caseNote, msg, caseSysOut))

            print("insert batch result...")
            resultXmlName = resultXmlName.replace('\\', '/')
            reportUrl = '%s/%s/log.html' % (self.ftpServer, os.path.dirname(resultXmlName).split('/')[-1])
            print("@@@@@@@@@@View the detail report:%s" % reportUrl)
            data = {'dbName': 'ArReport', 'fields': ['report_id'],
                    'data': [{'project_id': projectId, 'project_key': pkgName, 'report_name': projectName,
                              'report_type': subjectEnv, 'case_number': passCount + failCount, 'pass_field': passCount,
                              'fail': failCount, 'error': 0, 'skip': 0, 'run_time': sumTime, 'start_time': startTime,
                              'create_time': com.currentDateTime('%Y-%m-%d %H:%M:%S'),
                              'update_time': com.currentDateTime('%Y-%m-%d %H:%M:%S'),
                              'report_category': report_category, 'report_url': reportUrl}]}
            reportId = self.sendAddInfoApiRequest(data)[0]['report_id']

            casesList = []
            for case in caseRsList:
                casesList.append(
                    {'report_id': reportId, 'case_module': case[0], 'case_name': case[1], 'result': case[2],
                     'run_time': case[3], 'message': case[4], 'detail': case[5], 'system_out': case[6]})
            try:
                data = {'dbName': 'ArResult', 'data': casesList}
                self.sendAddInfoApiRequest(data)
            except Exception as e:
                print("@@@@@@@@@@@logUILogToDb:%s" % str(e))
        except Exception as e:
            print(e.message)
            return None, None
        return testRsFlag, reportId, reportUrl

    def apiLogToDb(self, reportId):
        '''
        through the way of the api to append api log to db
        :param reportId: the id of report on qa platform
        :return: None
        '''
        apiLogFile = 'data/api_log.txt'
        if os.path.exists(globalVariables.tempLogDir):
            apiLogFile = globalVariables.tempLogDir + '/api_log.txt'
        if os.path.exists(apiLogFile):
            fEnv = open(apiLogFile, 'r')
            apiLogs = fEnv.readlines()
            apiList = []
            tempApiList = []
            for line in apiLogs:
                values = line.replace('\n', '').strip('"').split('$')
                url = urllib.parse.unquote(values[3])
                iStart = url.find('/')
                if iStart > 0:
                    host = url[:iStart]
                    url = url[iStart:]
                else:
                    host = url
                    url = ''
                itemValue = {'case_module': values[1], 'case_name': values[2], 'host': host.strip(),
                             'url': url.strip(), 'method': values[4], 'report_id': reportId}
                if not itemValue in tempApiList:
                    tempValue = copy.copy(itemValue)
                    tempApiList.append((tempValue))
                    itemValue['date_time'] = values[0]
                    apiList.append(itemValue)
            data = {'dbName': 'ArApilog', 'data': apiList, 'forceUpdate': False}
            self.sendAddInfoApiRequest(data)

    def apiLogToDbV1(self, reportId):
        '''
        through the way of the sql to append api log to db
        :param reportId: the id of report on qa platform
        :return: None
        '''
        apiLogFile = 'data/api_log.txt'
        if os.path.exists(apiLogFile):
            fEnv = open(apiLogFile, 'r')
            apiLogs = fEnv.readlines()
            for line in apiLogs:
                values = line.replace('\n', '').strip('"').split('$')
                querySelect = '''select count(*) from `AR.ApiLog` where case_module = '%s' and case_name='%s' and
                                report_id=%d and url='%s'
                                ''' % (values[1], values[2], reportId, urllib.parse.unquote(values[3]))
                count = self.sendQaClubApiRequest(querySelect)[0]['count(*)']
                if count < 1:
                    insertQuery = '''insert into `AR.ApiLog`(date_time,case_module,case_name,url,method,report_id)
                      values('%s','%s','%s','%s','%s',%d)''' % (
                        values[0], values[1], values[2], urllib.parse.unquote(values[3]), values[4], reportId)
                    self.sendQaClubApiRequest(insertQuery, 1)


def main():
    reportId = 0
    runType = "api"
    parseLog = ParseAutoTestLog()
    htmlReport = HtmlReport()
    mail = Mail()
    args = sys.argv
    if len(args) > 1 and args[1] in ["-h", "-help", "/h", "/?"]:
        parseLog.showHelpInfo()
        return
    else:
        if not parseLog.loadCommandParas(args): return

    if globalVariables.tempLogDir == "" and os.path.exists("tempLogDir.txt"):
        f = open("tempLogDir.txt", 'r')
        globalVariables.tempLogDir = f.readline()
        f.close()
        os.remove("tempLogDir.txt")

    try:
        parseLog.projectKey = "testUIClassForPC"
        resultXmlName = "nosetests.xml"
        htmlResult = "nosetests.html"
        onlyUploadLog = False
        totalResult = "pass"
        run_env = ""
        if 'rFile' in parseLog.cmdParas:
            resultXmlName = parseLog.cmdParas['rFile']
        if 'projectKey' in parseLog.cmdParas:
            parseLog.projectKey = parseLog.cmdParas['projectKey']
            htmlReport.interfaceTitle = parseLog.cmdParas['projectKey']
        if 'reportCategory' in parseLog.cmdParas:
            parseLog.report_category = int(parseLog.cmdParas['reportCategory'])
        if 'runEnv' in parseLog.cmdParas:
            run_env = int(parseLog.cmdParas['runEnv'])
            mail.run_env = run_env
        if 'mailSubject' in parseLog.cmdParas:
            mail.mailSubject = parseLog.cmdParas['mailSubject']
        if 'mailSender' in parseLog.cmdParas:
            mail.mailSender = parseLog.cmdParas['mailSender']
        if 'mailList' in parseLog.cmdParas:
            mail.mailReceivers = parseLog.cmdParas['mailList']
        if 'runType' in parseLog.cmdParas:
            runType = parseLog.cmdParas['runType']

        if 'caseFilePath' in parseLog.cmdParas and parseLog.cmdParas['caseFilePath'] != []:
            parseLog.insertCaseDescToDb(parseLog.cmdParas['caseFilePath'].split(','))

        cwDir = os.getcwd()
        if os.path.exists(os.path.join(globalVariables.tempLogDir, resultXmlName)):
            resultXmlName = os.path.join(globalVariables.tempLogDir, resultXmlName)
        else:
            resultXmlName = os.path.join(cwDir, resultXmlName)
        if os.path.exists(os.path.join(globalVariables.tempLogDir, htmlResult)):
            htmlResult = os.path.join(globalVariables.tempLogDir, htmlResult)

        reportUrl = ""
        try:
            if runType == "api":
                if os.path.exists(htmlResult):
                    ftp.connect()
                    ftpLoc = 'api_%s' % (time.strftime("%Y%m%d_%H%M%S", time.localtime()))
                    if not ftp._is_ftp_dir('/autoLog/%s' % ftpLoc):
                        ftp.conn.mkd('/autoLog/%s' % ftpLoc)
                    ftp.put_file(htmlResult, '/autoLog/%s/nosetests.html' % ftpLoc)
                    ftp.close()
                    reportUrl = "http://xxx.xxx.xx.xxx/%s/nosetests.html" % ftpLoc
            else:
                source = cwDir + "/testlog/ui"
                print("@@@@@@@@@%s" % source)
                ftpfolder = "ui_nose_%s" % time.strftime("%Y%m%d_%H%M%S", time.localtime())
                dest = os.path.join("/autoLog", ftpfolder)
                reportUrl = '%s/%s/ui/nosetests.html' % (parseLog.ftpServer, ftpfolder)
                ftp.connect()
                ftp.put_dir(source, dest)
                ftp.close()
            ftpSuccess = True
        except:
            ftpSuccess = False

        fileType = resultXmlName.split('.')[-1]
        if str(onlyUploadLog).lower() == "false":
            parseLog.reportUrl = reportUrl
            if fileType == 'xml':
                testRsFlag, reportId = parseLog.logToDb(resultXmlName, run_env)
            elif fileType.find('xls') > -1:
                testRsFlag, reportId = parseLog.logExcelResultToDb(resultXmlName, run_env)
            if runType == "api":
                print("@@@@Start to add api log...")
                # parseLog.apiLogToDbV1(reportId)
                parseLog.apiLogToDb(reportId)
            print("http://qa.lb.com/report/apiReport/detail/" + str(reportId))

        # Send mail
        if "sendMail" in parseLog.cmdParas and parseLog.cmdParas["sendMail"].lower() == "true":
            # print "@@@@Start to analyze result file..."
            # if fileType == 'xml':
            #     casesSummary, casesResult = htmlReport.analyzeResultFile(resultXmlName)
            # elif fileType.find('xls') > -1:
            #     casesSummary, casesResult = htmlReport.analyzeResultExcelFile(resultXmlName)
            # htmlReport.reportUrl = reportUrl
            # print "@@@@Start to write result to html file..."
            showUrlInMail = True
            if 'showUrlInMail' in parseLog.cmdParas and parseLog.cmdParas["showUrlInMail"].lower() == "false":
                showUrlInMail = False
            # totalResult = htmlReport.writeTestDataToHtml(casesSummary, casesResult, "test.html", showUrlInMail,
            #                                              showDetailLink=ftpSuccess)
            # if parseLog.cmdParas.has_key('mailProjectKey'):
            #     casesSummary["interfaceTitle"] = parseLog.cmdParas['mailProjectKey']
            # print "@@@@Start to send out mail..."
            # mail.sendMail(casesSummary, totalResult)
            # os.remove("test.html")
            # if os.path.exists(os.path.join(globalVariables.tempLogDir)):
            #     shutil.rmtree(os.path.join(globalVariables.tempLogDir))
            #     # os.remove(resultXmlName)
            headers = {'Content-Type': 'application/json', 'User-Agent': '1qazXSW@Auto'}
            url = "http://xxx.xxx.xx.xxx:8000/ci/api/sendMail"
            try:
                import requests, json

                data = {'report_id': reportId, 'showUrlInMail': showUrlInMail}
                if 'mailSubject' in parseLog.cmdParas:
                    data['mailSubject'] = parseLog.cmdParas['mailSubject'].decode('gbk')
                if 'mailSender' in parseLog.cmdParas:
                    data['mailSender'] = parseLog.cmdParas['mailSender']
                if 'mailList' in parseLog.cmdParas:
                    data['mailReceivers'] = parseLog.cmdParas['mailList']

                resp = requests.post(url, json.dumps(data, ensure_ascii=False), headers=headers)
                resp = json.loads(resp.content)
                print(("sendMailApiRequest: %s" % resp['message']))
            except Exception as e:
                print(("sendMailApiRequest: %s" % str(e)))

        print("@@@@@@@@@@View the detail report:%s" % reportUrl)

        if totalResult != "pass":
            sys.exit(1)
    except Exception as e:
        msg = traceback.format_exc()
        print(msg)
        # print "@@@@@@@@@Exception:%s" % str(e)
        sys.exit(1)


if __name__ == "__main__":
    os.chdir(r'C:\Git\automation_api\XXX')
    main()

    # parseLog = ParseAutoTestLog()
    # parseLog.projectKey = "testApiWebSite"
    # caseFilePath = []
    # resultXmlName = "nosetests.xml"
    # parseLog.report_category = 0
    # run_env = ""
    # if len(sys.argv) > 1:
    #     resultXmlName = sys.argv[1]
    # if len(sys.argv) > 2:
    #     parseLog.projectKey = sys.argv[2]
    # if len(sys.argv) > 3 and sys.argv[3] != "":
    #     for item in sys.argv[3].split(','):
    #         caseFilePath.append(os.getcwd() + '/' + item)
    # if len(sys.argv) > 4:
    #     parseLog.report_category = int(sys.argv[4])
    # if len(sys.argv) > 5:
    #     run_env = int(sys.argv[5])
    #
    # reload(sys)
    # sys.setdefaultencoding('utf-8')
    # try:
    #     # caseFilePath = [r'E:\Work\home\automation_api\XXX\testcase\testBaseLine\testApiForMetaData']
    #     if caseFilePath != []:
    #         parseLog.insertCaseDescToDb(caseFilePath)
    #     fileType = resultXmlName.split('.')[-1]
    #     if fileType == 'xml':
    #         testRsFlag, reportId = parseLog.logToDb(resultXmlName, run_env)
    #     elif fileType.find('xls') > -1:
    #         testRsFlag, reportId = parseLog.logExcelResultToDb(resultXmlName)
    #     print testRsFlag
    #     print "http://lbxx.com.com/report/apiReport/detail/" + str(reportId)
    # except:
    #     sys.exit(1)
